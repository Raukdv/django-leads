from django import forms
from django.core.validators import FileExtensionValidator
from django.utils.html import mark_safe

#Tools
import os
import csv
import codecs

#Openxlpy
from openpyxl import load_workbook, Workbook

class FileUploadValidatorForm(forms.Form):
    ALLOWED_EXTENSIONS = ['csv', 'xlsx', 'xls']

    file = forms.FileField(required=True, label=("file"), help_text="Just CSV/XLSX or XLS will be processed.", 
            widget=forms.FileInput(attrs={
                'class': 'hidden'
                }
            ),
            validators=[
                FileExtensionValidator(allowed_extensions=ALLOWED_EXTENSIONS)
            ]
        )

    def clean_file(self):
        file = self.cleaned_data['file']
        file_extension = os.path.splitext(file.name)[1]
        file_clean = file_extension[1:].lower()
        getattr(self, f'clean_file_{file_clean}')(file=file)
        
        return file
    
    def clean_file_csv(self, file):
        try:
            reader = csv.DictReader(codecs.iterdecode(file, 'utf-8'))
        except UnicodeDecodeError:
            raise forms.ValidationError(("Invalid file content. Contains invalid characters.")
            )

        index = 0
        self.form_list = []
        for row in reader:
            index += 1
            validated_row = self.validate_row(row, index)
            if validated_row:
                self.form_list.append(validated_row)
        return file
    
    def clean_file_xlsx(self, file):
        wb = load_workbook(file)
        ws = wb.worksheets[0]
        index = 0
        header = None
        self.form_list = []
        for row in ws:
            index += 1
            
            row = [cel.value for cel in row]
            if(index == 1):
                header = row
                continue

            row = dict(zip(header, row))

            if not any(row.values()):
                continue

            row = self.validate_row(row, index)
            if row:
                self.form_list.append(row)

        return file

    def clean_file_xls(self, file):
        wb = load_workbook(file)
        ws = wb.worksheets[0]
        index = 0
        header = None
        self.form_list = []
        for row in ws:
            index += 1
            
            row = [cel.value for cel in row]
            if(index == 1):
                header = row
                continue

            row = dict(zip(header, row))

            if not any(row.values()):
                continue

            row = self.validate_row(row, index)
            if row:
                self.form_list.append(row)

        return file
    
    def validate_row(self, row, index=0):
        form = self.get_form(data=row)
        if not form.is_valid():
            message_dict = form.errors
            errors = []
            for field in message_dict:
                errors.append('%(field)s: %(error)s' % dict(field=field, error=' '.join(message_dict[field])
                    )
                )
            message = ("Line %(index)d:<br>%(content)s") % dict(index=index, content='<br>'.join(errors))
            self.add_error('file', mark_safe(message))

        return form

    def get_form(self, form_class=None, **kwargs):
        form_class = form_class or self.get_form_class()
        return form_class(**kwargs)

    def get_form_class(self):
        return self.form_class